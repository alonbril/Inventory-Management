from flask import Flask, render_template, request, redirect, url_for, flash, g, session, send_from_directory, jsonify
from database import init_db, get_db, close_db, get_db_path, verify_database_structure
from datetime import datetime, date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename

import os
import sys
import logging


# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Define DATABASE path before app creation
DATABASE = get_db_path()

app = Flask(__name__,
            template_folder=resource_path('templates'),
            static_folder=resource_path('static'))

app.secret_key = 'dev'

# Configure upload folder
UPLOAD_FOLDER = resource_path('uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_pagination_params():
    """Get pagination parameters from request"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page
    return page, per_page, offset


def get_sort_params():
    """Get sorting parameters from request"""
    sort_by = request.args.get('sort_by', 'id')
    sort_order = request.args.get('sort_order', 'desc')

    # Validate sort parameters
    allowed_sort_fields = {
        'id': 'id',
        'green_number': 'green_number',
        'date': 'created_at'  # for inventory items
    }

    if sort_by not in allowed_sort_fields:
        sort_by = 'id'

    sort_field = allowed_sort_fields[sort_by]
    sort_direction = 'DESC' if sort_order == 'desc' else 'ASC'

    return sort_field, sort_direction

@app.template_filter('max')
def max_filter(a, b):
    return max(a, b)

@app.template_filter('min')
def min_filter(a, b):
    return min(a, b)


# Error handling
@app.errorhandler(Exception)
def handle_error(e):
    logger.error(f"Unhandled error: {str(e)}", exc_info=True)
    return render_template('error.html', error=str(e)), 500

@app.before_request
def before_request():
    """Establish database connection before each request"""
    try:
        get_db()
    except Exception as e:
        logger.error(f"Database connection error: {str(e)}", exc_info=True)
        return render_template('error.html', error="Database connection failed"), 500

app.teardown_appcontext(close_db)

def init_app():
    """Initialize the application"""
    try:
        with app.app_context():
            verify_database_structure()
            init_db()
            sync_inventory_status()
    except Exception as e:
        logger.error(f"Initialization error: {str(e)}", exc_info=True)
        raise


@app.route('/')
def index():
    db = get_db()
    cursor = db.cursor()

    # Get pagination and sorting parameters
    page, per_page, offset = get_pagination_params()
    sort_field, sort_direction = get_sort_params()

    # Get search query from URL parameters
    search_query = request.args.get('search', '').strip()

    # First, get the overdue loan green numbers
    cursor.execute('''
        SELECT DISTINCT l.green_number 
        FROM loans l 
        WHERE l.status = 'active' 
        AND CAST((JULIANDAY('now') - JULIANDAY(l.loan_date)) AS INTEGER) > 7
    ''')
    overdue_items = cursor.fetchall()
    overdue_green_numbers = set(str(item['green_number']) for item in overdue_items)

    # Base query for counting total items
    count_query = 'SELECT COUNT(*) as total FROM inventory'

    # Base query for fetching items
    item_query = '''
        SELECT * FROM inventory
        {where_clause}
        ORDER BY {sort_field} {sort_direction}
        LIMIT ? OFFSET ?
    '''

    # Add search conditions if search_query exists
    where_clause = ''
    query_params = []

    if search_query:
        where_clause = '''
            WHERE name LIKE ? 
            OR green_number LIKE ? 
            OR category LIKE ?
            OR status LIKE ?
        '''
        query_params = [f'%{search_query}%'] * 4

    # Get total count
    if search_query:
        cursor.execute(count_query + ' ' + where_clause, query_params)
    else:
        cursor.execute(count_query)

    total_items = cursor.fetchone()['total']
    total_pages = (total_items + per_page - 1) // per_page

    # Get items for current page
    query = item_query.format(
        where_clause=where_clause,
        sort_field=sort_field,
        sort_direction=sort_direction
    )

    cursor.execute(query, query_params + [per_page, offset])
    items = [dict(row) for row in cursor.fetchall()]

    # Add is_overdue flag to each item
    for item in items:
        item['is_overdue'] = str(item['green_number']) in overdue_green_numbers

    return render_template('index.html',
                           items=items,
                           search_query=search_query,
                           current_page=page,
                           total_pages=total_pages,
                           sort_by=request.args.get('sort_by', 'id'),
                           sort_order=request.args.get('sort_order', 'desc'))


@app.route('/add', methods=['GET', 'POST'])
def add_item():
    if request.method == 'POST':
        try:
            name = request.form['name']
            quantity = int(request.form['quantity'])
            green_number = int(request.form['green_number'])
            category = request.form['category']
            status = request.form.get('status', 'no')

            db = get_db()
            cursor = db.cursor()
            cursor.execute(
                'INSERT INTO inventory (name, quantity, green_number, category, status) VALUES (?, ?, ?, ?, ?)',
                (name, quantity, green_number, category, status)
            )
            db.commit()
            flash('Item added successfully!', 'success')
        except Exception as e:
            flash(f'Error adding item: {str(e)}', 'error')
            db.rollback()
        return redirect(url_for('index'))

    return render_template('add.html')


@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_item(id):
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        try:
            name = request.form['name']
            quantity = int(request.form['quantity'])
            green_number = int(request.form['green_number'])
            category = request.form['category']
            status = request.form.get('status', 'no')

            cursor.execute('''
                UPDATE inventory 
                SET name = ?, quantity = ?, green_number = ?, category = ?, status = ?
                WHERE id = ?
            ''', (name, quantity, green_number, category, status, id))
            db.commit()
            flash('Item updated successfully!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Error updating item: {str(e)}', 'error')
            db.rollback()

    cursor.execute('SELECT * FROM inventory WHERE id = ?', (id,))
    item = cursor.fetchone()
    return render_template('edit.html', item=item)


@app.route('/loans')
def loans():
    db = get_db()
    cursor = db.cursor()

    # Get all active loans with calculated days active
    cursor.execute('''
        SELECT *, 
        CAST((JULIANDAY('now') - JULIANDAY(loan_date)) AS INTEGER) as days_active 
        FROM loans 
        WHERE status = 'active' 
        ORDER BY borrower_name, loan_date DESC
    ''')

    loans = [dict(loan) for loan in cursor.fetchall()]

    # Add is_overdue flag to each loan
    for loan in loans:
        loan['is_overdue'] = loan['days_active'] > 7

    return render_template('loans.html', loans=loans)


@app.route('/add_loan', methods=['GET', 'POST'])
def add_loan():
    try:
        db = get_db()
        cursor = db.cursor()

        if request.method == 'POST':
            try:
                borrower_name = request.form['borrower_name']
                green_numbers = request.form.getlist('green_numbers[]')
                equipment = request.form.getlist('equipment[]')
                equipment_quantities = request.form.getlist('equipment_quantity[]')
                loan_date = request.form['loan_date']
                signature = request.form.get('signature', '')

                if not signature:
                    flash('Signature is required!', 'error')
                    return render_template('add_loan.html',
                                         form_data=request.form,
                                         available_items=get_available_items(),
                                         cart_templates=get_cart_templates())

                if not green_numbers or not any(green_numbers):
                    flash('Please select at least one item!', 'error')
                    return render_template('add_loan.html',
                                         form_data=request.form,
                                         available_items=get_available_items(),
                                         cart_templates=get_cart_templates())

                # Start a transaction
                cursor.execute('BEGIN TRANSACTION')

                # Process each green number as a separate loan
                for green_number in green_numbers:
                    if not green_number:  # Skip empty selections
                        continue

                    # Check if green number exists and is available
                    cursor.execute('SELECT * FROM inventory WHERE green_number = ?', (green_number,))
                    inventory_item = cursor.fetchone()

                    if inventory_item is None:
                        raise Exception(f'Green number {green_number} does not exist in inventory!')

                    cursor.execute('''
                        SELECT * FROM loans 
                        WHERE green_number = ? 
                        AND status = 'active'
                    ''', (green_number,))
                    existing_loan = cursor.fetchone()

                    if existing_loan:
                        raise Exception(f'Green number {green_number} is currently on loan!')

                    # Create the loan record
                    cursor.execute('''
                        INSERT INTO loans 
                        (borrower_name, item_name, green_number, loan_date, signature, status) 
                        VALUES (?, ?, ?, ?, ?, 'active')
                    ''', (borrower_name, inventory_item['name'], green_number, loan_date, signature))

                    # Update inventory item status to 'yes'
                    cursor.execute('''
                        UPDATE inventory
                        SET status = 'yes'
                        WHERE green_number = ?
                    ''', (green_number,))

                    loan_id = cursor.lastrowid

                    # Add equipment for this loan
                    for equip, qty in zip(equipment, equipment_quantities):
                        if equip:  # Only process if equipment is selected
                            cursor.execute('''
                                INSERT INTO loans_equipment 
                                (loan_id, equipment_type, quantity) 
                                VALUES (?, ?, ?)
                            ''', (loan_id, equip, qty))

                # Commit the transaction
                cursor.execute('COMMIT')
                flash('Loan(s) added successfully!', 'success')
                return redirect(url_for('loans'))

            except Exception as e:
                # If anything goes wrong, rollback the transaction
                cursor.execute('ROLLBACK')
                flash(f'Error adding loan: {str(e)}', 'error')
                return render_template('add_loan.html',
                                     form_data=request.form,
                                     available_items=get_available_items(),
                                     cart_templates=get_cart_templates())

        # For GET requests
        return render_template('add_loan.html',
                             available_items=get_available_items(),
                             cart_templates=get_cart_templates())

    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('loans'))

# Add this helper function if not already present
def get_cart_templates():
    db = get_db()
    cursor = db.cursor()
    cursor.execute('SELECT id, name FROM cart_templates ORDER BY name')
    return cursor.fetchall()


def is_green_number_in_use(green_number, exclude_template_id=None):
    """
    Check if a green number is already used in any other template

    Args:
        green_number: The green number to check
        exclude_template_id: Optional template ID to exclude from the check (used during editing)

    Returns:
        bool: True if the green number is already in use, False otherwise
    """
    db = get_db()
    cursor = db.cursor()

    if exclude_template_id is not None:
        cursor.execute('''
            SELECT COUNT(*) as count 
            FROM cart_template_items cti 
            JOIN cart_templates ct ON cti.template_id = ct.id 
            WHERE cti.green_number = ? AND ct.id != ?
        ''', (green_number, exclude_template_id))
    else:
        cursor.execute('''
            SELECT COUNT(*) as count 
            FROM cart_template_items 
            WHERE green_number = ?
        ''', (green_number,))

    result = cursor.fetchone()
    return result['count'] > 0


# Add this helper function to get available items
def get_available_items():
    db = get_db()
    cursor = db.cursor()

    # Get items that are either not loaned or have been returned
    cursor.execute('''
        SELECT DISTINCT i.green_number, i.name 
        FROM inventory i
        LEFT JOIN loans l ON i.green_number = l.green_number AND l.status = 'active'
        WHERE l.id IS NULL
        ORDER BY i.green_number
    ''')

    return cursor.fetchall()

def get_active_loans():
    db = get_db()
    cursor = db.cursor()
    cursor.execute('''
        SELECT green_number, item_name, borrower_name, loan_date 
        FROM loans 
        WHERE status = 'active' 
        ORDER BY loan_date DESC
    ''')
    return cursor.fetchall()


@app.route('/return_loan/<int:id>')
def return_loan(id):
    try:
        db = get_db()
        cursor = db.cursor()

        # Start transaction
        cursor.execute('BEGIN TRANSACTION')

        # Get the green number before updating the loan
        cursor.execute('SELECT green_number FROM loans WHERE id = ?', (id,))
        loan = cursor.fetchone()

        if loan:
            # Update loan status to returned
            cursor.execute('''
                UPDATE loans 
                SET status = 'returned', return_date = ?
                WHERE id = ?
            ''', (date.today().isoformat(), id))

            # Check if this item has any other active loans
            cursor.execute('''
                SELECT COUNT(*) as active_count 
                FROM loans 
                WHERE green_number = ? AND status = 'active' AND id != ?
            ''', (loan['green_number'], id))

            active_count = cursor.fetchone()['active_count']

            # If no other active loans, update inventory status to 'no'
            if active_count == 0:
                cursor.execute('''
                    UPDATE inventory 
                    SET status = 'no'
                    WHERE green_number = ?
                ''', (loan['green_number'],))

        # Commit transaction
        db.commit()
        flash('Loan marked as returned!', 'success')

    except Exception as e:
        flash(f'Error updating loan: {str(e)}', 'error')
        db.rollback()
    return redirect(url_for('loans'))


@app.route('/get_template_items/<int:template_id>')
def get_template_items(template_id):
    try:
        db = get_db()
        cursor = db.cursor()

        # Get all items for the template
        cursor.execute('''
            SELECT cti.green_number, i.name 
            FROM cart_template_items cti 
            JOIN inventory i ON cti.green_number = i.green_number 
            WHERE cti.template_id = ?
        ''', (template_id,))

        items = cursor.fetchall()
        return jsonify([dict(item) for item in items])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/delete/<int:id>')
def delete_item(id):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM inventory WHERE id = ?', (id,))
        db.commit()
        flash('Item deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting item: {str(e)}', 'error')
        db.rollback()
    return redirect(url_for('index'))


@app.route('/loans_history')
def loans_history():
    db = get_db()
    cursor = db.cursor()

    # Get pagination and sorting parameters
    page, per_page, offset = get_pagination_params()
    sort_field, sort_direction = get_sort_params()

    # Count total returned loans
    cursor.execute('''
        SELECT COUNT(*) as total 
        FROM loans 
        WHERE status = 'returned'
    ''')
    total_items = cursor.fetchone()['total']
    total_pages = (total_items + per_page - 1) // per_page

    # Get loans for current page
    cursor.execute('''
        SELECT * FROM loans 
        WHERE status = 'returned' 
        ORDER BY {} {}
        LIMIT ? OFFSET ?
    '''.format(sort_field, sort_direction), [per_page, offset])

    loans = cursor.fetchall()

    return render_template('loans_history.html',
                           loans=loans,
                           current_page=page,
                           total_pages=total_pages,
                           sort_by=request.args.get('sort_by', 'id'),
                           sort_order=request.args.get('sort_order', 'desc'))



UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False

# Create uploads folder if it doesn't exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def create_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    # Define headers
    headers = ['name', 'quantity', 'green_number', 'category', 'status']
    example_data = ['Example Item', 5, 1001, 'Electronics', 'no']

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply headers and styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Add example row
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 30  # name
    ws.column_dimensions['B'].width = 10  # quantity
    ws.column_dimensions['C'].width = 15  # green_number
    ws.column_dimensions['D'].width = 20  # category
    ws.column_dimensions['E'].width = 10  # status

    # Create templates directory if it doesn't exist
    template_dir = os.path.join(app.static_folder, 'templates')
    os.makedirs(template_dir, exist_ok=True)

    # Save template
    template_path = os.path.join(template_dir, 'inventory_template.xlsx')
    wb.save(template_path)
    return template_path


# Add this route for template download
@app.route('/download_template')
def download_template():
    template_path = create_excel_template()
    directory = os.path.dirname(template_path)
    filename = os.path.basename(template_path)
    return send_from_directory(directory, filename, as_attachment=True)


# Update your import_inventory route to include template creation
@app.route('/import_inventory', methods=['GET', 'POST'])
def import_inventory():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('index'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('index'))

        if file and allowed_file(file.filename):
            try:
                # Read Excel file
                df = pd.read_excel(file)

                # Verify required columns
                required_columns = {'name', 'quantity', 'green_number', 'category', 'status'}
                if not required_columns.issubset(df.columns):
                    flash('Excel file must contain columns: name, quantity, green_number, category, status', 'error')
                    return redirect(url_for('index'))

                # Process each row
                db = get_db()
                cursor = db.cursor()
                success_count = 0
                error_count = 0

                for index, row in df.iterrows():
                    try:
                        # Check if item with green_number already exists
                        cursor.execute('SELECT id FROM inventory WHERE green_number = ?', (row['green_number'],))
                        existing_item = cursor.fetchone()

                        if existing_item:
                            # Update existing item
                            cursor.execute('''
                                UPDATE inventory 
                                SET name = ?, quantity = ?, category = ?, status = ?
                                WHERE green_number = ?
                            ''', (row['name'], row['quantity'], row['category'],
                                  row['status'], row['green_number']))
                        else:
                            # Insert new item
                            cursor.execute('''
                                INSERT INTO inventory (name, quantity, green_number, category, status)
                                VALUES (?, ?, ?, ?, ?)
                            ''', (row['name'], row['quantity'], row['green_number'],
                                  row['category'], row['status']))
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error processing row {index}: {str(e)}")
                        continue

                db.commit()
                flash(f'Successfully imported {success_count} items. {error_count} errors.', 'success')

            except Exception as e:
                flash(f'Error processing file: {str(e)}', 'error')
                logger.error(f"Import error: {str(e)}")

            return redirect(url_for('index'))

        flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('index'))

    # Create template on first access
    template_path = os.path.join(app.static_folder, 'templates', 'inventory_template.xlsx')
    if not os.path.exists(template_path):
        create_excel_template()

    return render_template('import_inventory.html')

# Add this function at the end of your app.py, just before the if __name__ == '__main__': line

def sync_inventory_status():
    try:
        db = get_db()
        cursor = db.cursor()

        # Start transaction
        cursor.execute('BEGIN TRANSACTION')

        # First, set all items to 'no'
        cursor.execute('UPDATE inventory SET status = ?', ('no',))

        # Then set items with active loans to 'yes'
        cursor.execute('''
            UPDATE inventory 
            SET status = 'yes'
            WHERE green_number IN (
                SELECT DISTINCT green_number 
                FROM loans 
                WHERE status = 'active'
            )
        ''')

        db.commit()
        return True
    except Exception as e:
        db.rollback()
        print(f"Error syncing inventory status: {str(e)}")
        return False


@app.route('/extend_loan/<int:id>')
def extend_loan(id):
    try:
        db = get_db()
        cursor = db.cursor()

        # Get current loan details
        cursor.execute('''
            SELECT loan_date, 
            CAST((JULIANDAY('now') - JULIANDAY(loan_date)) AS INTEGER) as days_active 
            FROM loans 
            WHERE id = ?
        ''', (id,))
        loan = cursor.fetchone()

        if loan and loan['days_active'] > 7:
            # Calculate new loan date (current loan date + 7 days)
            cursor.execute('''
                UPDATE loans 
                SET loan_date = date(loan_date, '+7 days')
                WHERE id = ?
            ''', (id,))
            db.commit()
            flash('Loan extended for one week!', 'success')
        else:
            flash('Loan cannot be extended!', 'error')

    except Exception as e:
        flash(f'Error extending loan: {str(e)}', 'error')
        db.rollback()
    return redirect(url_for('loans'))


@app.route('/loan/<int:id>')
def loan_details(id):
    try:
        db = get_db()
        cursor = db.cursor()

        # Store the current page in session for the back button
        session['previous_page'] = request.referrer

        # Get loan details
        cursor.execute('''
            SELECT *, 
            CAST((JULIANDAY('now') - JULIANDAY(loan_date)) AS INTEGER) as days_active 
            FROM loans 
            WHERE id = ?
        ''', (id,))
        loan = cursor.fetchone()

        if not loan:
            flash('Loan not found!', 'error')
            return redirect(url_for('loans'))

        # Get equipment details
        cursor.execute('''
            SELECT equipment_type, quantity 
            FROM loans_equipment 
            WHERE loan_id = ?
        ''', (id,))
        equipment = cursor.fetchall()

        return render_template('loan_details.html',
                               loan=loan,
                               equipment=equipment,
                               days_active=loan['days_active'])

    except Exception as e:
        flash(f'Error loading loan details: {str(e)}', 'error')
        return redirect(url_for('loans'))


@app.route('/back')
def back_to_previous():
    previous_page = session.get('previous_page')
    if previous_page:
        return redirect(previous_page)
    return redirect(url_for('loans'))

# Optional: Add the route to manually trigger the sync
@app.route('/sync_inventory_status')
def sync_inventory():
    if sync_inventory_status():
        flash('Inventory status synchronized successfully!', 'success')
    else:
        flash('Error synchronizing inventory status!', 'error')
    return redirect(url_for('index'))


@app.route('/toner')
def toner_management():
    db = get_db()
    cursor = db.cursor()

    # Get search query from URL parameters
    search_query = request.args.get('search', '').strip()

    if search_query:
        cursor.execute('''
            SELECT * FROM toner_inventory 
            WHERE name LIKE ? 
            OR printer LIKE ? 
            OR color LIKE ?
            ORDER BY id DESC
        ''', (f'%{search_query}%', f'%{search_query}%', f'%{search_query}%'))
    else:
        cursor.execute('SELECT * FROM toner_inventory ORDER BY id DESC')

    toners = cursor.fetchall()
    return render_template('toner.html', toners=toners, search_query=search_query)


@app.route('/add_toner', methods=['GET', 'POST'])
def add_toner():
    if request.method == 'POST':
        try:
            name = request.form['name']
            printer = request.form['printer']
            bk_toner = request.form['bk_toner']
            color = request.form['color']
            inventory = int(request.form['inventory'])

            db = get_db()
            cursor = db.cursor()
            cursor.execute(
                'INSERT INTO toner_inventory (name, printer, bk_toner, color, inventory) VALUES (?, ?, ?, ?, ?)',
                (name, printer, bk_toner, color, inventory)
            )
            db.commit()
            flash('Toner added successfully!', 'success')
        except Exception as e:
            flash(f'Error adding toner: {str(e)}', 'error')
            db.rollback()
        return redirect(url_for('toner_management'))

    return render_template('add_toner.html')


@app.route('/edit_toner/<int:id>', methods=['GET', 'POST'])
def edit_toner(id):
    db = get_db()
    cursor = db.cursor()

    if request.method == 'POST':
        try:
            name = request.form['name']
            printer = request.form['printer']
            bk_toner = request.form['bk_toner']
            color = request.form['color']
            inventory = int(request.form['inventory'])

            cursor.execute('''
                UPDATE toner_inventory 
                SET name = ?, printer = ?, bk_toner = ?, color = ?, inventory = ?
                WHERE id = ?
            ''', (name, printer, bk_toner, color, inventory, id))
            db.commit()
            flash('Toner updated successfully!', 'success')
            return redirect(url_for('toner_management'))
        except Exception as e:
            flash(f'Error updating toner: {str(e)}', 'error')
            db.rollback()

    cursor.execute('SELECT * FROM toner_inventory WHERE id = ?', (id,))
    toner = cursor.fetchone()
    return render_template('edit_toner.html', toner=toner)


@app.route('/delete_toner/<int:id>')
def delete_toner(id):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM toner_inventory WHERE id = ?', (id,))
        db.commit()
        flash('Toner deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting toner: {str(e)}', 'error')
        db.rollback()
    return redirect(url_for('toner_management'))


@app.route('/cart_templates')
def cart_templates():
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute('''
            SELECT ct.*, COUNT(cti.id) as item_count 
            FROM cart_templates ct 
            LEFT JOIN cart_template_items cti ON ct.id = cti.template_id 
            GROUP BY ct.id
            ORDER BY ct.name
        ''')
        templates = cursor.fetchall()
        return render_template('cart_templates.html', templates=templates)
    except Exception as e:
        flash(f'Error loading templates: {str(e)}', 'error')
        return redirect(url_for('index'))


@app.route('/add_cart_template', methods=['GET', 'POST'])
def add_cart_template():
    try:
        db = get_db()
        cursor = db.cursor()

        if request.method == 'POST':
            name = request.form['name']
            green_numbers = request.form.getlist('green_numbers[]')

            if not green_numbers:
                flash('Please select at least one item for the template', 'error')
                return redirect(url_for('add_cart_template'))

            # Check for duplicate green numbers
            duplicate_numbers = []
            for green_number in green_numbers:
                if is_green_number_in_use(green_number):
                    cursor.execute('SELECT ct.name FROM cart_templates ct JOIN cart_template_items cti ON ct.id = cti.template_id WHERE cti.green_number = ?', (green_number,))
                    template = cursor.fetchone()
                    duplicate_numbers.append(f"Green number {green_number} is already in template '{template['name']}'")

            if duplicate_numbers:
                flash('Error: \n' + '\n'.join(duplicate_numbers), 'error')
                cursor.execute('SELECT green_number, name FROM inventory ORDER BY green_number')
                inventory_items = cursor.fetchall()
                return render_template('add_edit_cart_template.html',
                                    inventory_items=inventory_items,
                                    template=None,
                                    selected_items=green_numbers)

            cursor.execute('INSERT INTO cart_templates (name) VALUES (?)', (name,))
            template_id = cursor.lastrowid

            for green_number in green_numbers:
                cursor.execute('''
                    INSERT INTO cart_template_items (template_id, green_number) 
                    VALUES (?, ?)
                ''', (template_id, green_number))

            db.commit()
            flash('Template created successfully!', 'success')
            return redirect(url_for('cart_templates'))

        # GET request - show form
        cursor.execute('''
            SELECT i.green_number, i.name 
            FROM inventory i 
            WHERE NOT EXISTS (
                SELECT 1 FROM cart_template_items cti 
                WHERE cti.green_number = i.green_number
            )
            ORDER BY i.green_number
        ''')
        inventory_items = cursor.fetchall()
        return render_template('add_edit_cart_template.html',
                            inventory_items=inventory_items,
                            template=None,
                            selected_items=[])

    except Exception as e:
        flash(f'Error creating template: {str(e)}', 'error')
        return redirect(url_for('cart_templates'))

# Update the edit_cart_template route
@app.route('/edit_cart_template/<int:id>', methods=['GET', 'POST'])
def edit_cart_template(id):
    try:
        db = get_db()
        cursor = db.cursor()

        if request.method == 'POST':
            name = request.form['name']
            green_numbers = request.form.getlist('green_numbers[]')

            if not green_numbers:
                flash('Please select at least one item for the template', 'error')
                return redirect(url_for('edit_cart_template', id=id))

            # Check for duplicate green numbers
            duplicate_numbers = []
            for green_number in green_numbers:
                if is_green_number_in_use(green_number, exclude_template_id=id):
                    cursor.execute('SELECT ct.name FROM cart_templates ct JOIN cart_template_items cti ON ct.id = cti.template_id WHERE cti.green_number = ?', (green_number,))
                    template = cursor.fetchone()
                    duplicate_numbers.append(f"Green number {green_number} is already in template '{template['name']}'")

            if duplicate_numbers:
                flash('Error: \n' + '\n'.join(duplicate_numbers), 'error')
                cursor.execute('SELECT green_number, name FROM inventory ORDER BY green_number')
                inventory_items = cursor.fetchall()
                cursor.execute('SELECT green_number FROM cart_template_items WHERE template_id = ?', (id,))
                selected_items = [item['green_number'] for item in cursor.fetchall()]
                return render_template('add_edit_cart_template.html',
                                    template={'id': id, 'name': name},
                                    inventory_items=inventory_items,
                                    selected_items=green_numbers)

            cursor.execute('UPDATE cart_templates SET name = ? WHERE id = ?', (name, id))
            cursor.execute('DELETE FROM cart_template_items WHERE template_id = ?', (id,))

            for green_number in green_numbers:
                cursor.execute('''
                    INSERT INTO cart_template_items (template_id, green_number) 
                    VALUES (?, ?)
                ''', (id, green_number))

            db.commit()
            flash('Template updated successfully!', 'success')
            return redirect(url_for('cart_templates'))

        # GET request - show form with existing data
        cursor.execute('SELECT * FROM cart_templates WHERE id = ?', (id,))
        template = cursor.fetchone()

        if not template:
            flash('Template not found', 'error')
            return redirect(url_for('cart_templates'))

        cursor.execute('SELECT green_number FROM cart_template_items WHERE template_id = ?', (id,))
        selected_items = [item['green_number'] for item in cursor.fetchall()]

        # Get available items plus currently selected items
        cursor.execute('''
            SELECT i.green_number, i.name 
            FROM inventory i 
            WHERE NOT EXISTS (
                SELECT 1 FROM cart_template_items cti 
                WHERE cti.green_number = i.green_number 
                AND cti.template_id != ?
            )
            ORDER BY i.green_number
        ''', (id,))
        inventory_items = cursor.fetchall()

        return render_template('add_edit_cart_template.html',
                            template=template,
                            inventory_items=inventory_items,
                            selected_items=selected_items)

    except Exception as e:
        flash(f'Error editing template: {str(e)}', 'error')
        return redirect(url_for('cart_templates'))

@app.route('/delete_cart_template/<int:id>')
def delete_cart_template(id):
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('DELETE FROM cart_templates WHERE id = ?', (id,))
        db.commit()
        flash('Template deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting template: {str(e)}', 'error')
    return redirect(url_for('cart_templates'))


# Add these new routes to your app.py

@app.route('/bulk_return', methods=['POST'])
def bulk_return():
    try:
        db = get_db()
        cursor = db.cursor()

        loan_ids = request.form.getlist('loan_ids[]')

        # Start transaction
        cursor.execute('BEGIN TRANSACTION')

        for loan_id in loan_ids:
            # Get the green number before updating the loan
            cursor.execute('SELECT green_number FROM loans WHERE id = ?', (loan_id,))
            loan = cursor.fetchone()

            if loan:
                # Update loan status to returned
                cursor.execute('''
                    UPDATE loans 
                    SET status = 'returned', return_date = ?
                    WHERE id = ?
                ''', (date.today().isoformat(), loan_id))

                # Check if this item has any other active loans
                cursor.execute('''
                    SELECT COUNT(*) as active_count 
                    FROM loans 
                    WHERE green_number = ? AND status = 'active' AND id != ?
                ''', (loan['green_number'], loan_id))

                active_count = cursor.fetchone()['active_count']

                # If no other active loans, update inventory status to 'no'
                if active_count == 0:
                    cursor.execute('''
                        UPDATE inventory 
                        SET status = 'no'
                        WHERE green_number = ?
                    ''', (loan['green_number'],))

        # Commit transaction
        db.commit()
        flash('Selected loans have been returned successfully!', 'success')

    except Exception as e:
        cursor.execute('ROLLBACK')
        flash(f'Error returning loans: {str(e)}', 'error')

    return redirect(url_for('loans'))


@app.route('/get_borrower_loans/<borrower_name>')
def get_borrower_loans(borrower_name):
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute('''
            SELECT id, green_number, item_name 
            FROM loans 
            WHERE borrower_name = ? AND status = 'active'
            ORDER BY loan_date DESC
        ''', (borrower_name,))

        loans = cursor.fetchall()
        return jsonify([dict(loan) for loan in loans])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


DATABASE = get_db_path()

if __name__ == '__main__':
    try:
        init_app()
        app.run(debug=True, host='0.0.0.0')
    except Exception as e:
        logger.error(f"Startup error: {str(e)}", exc_info=True)
        sys.exit(1)